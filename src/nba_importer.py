#from pathlib import Path
import pandas as pd
import xlwings as xw

# Define the current script folder and sibling folder
#current_folder = Path(__file__).parent  # Current script directory (src)
#target_folder = current_folder.parent / "dk-data"  # Sibling folder (dk-data)

# Define file paths
#excel_file = target_folder / "nba_dfs_optimized.xlsm"
#csv_file = target_folder / "DKSalaries.csv"
excel_file = "nba_dfs_optimized.xlsm"
csv_file = "DKEntries.csv"

app = xw.App(visible=False)  # Set to True to see the Excel window
wb = app.books.open(str(excel_file))


# Load the CSV data
data = pd.read_csv(csv_file)

# Define the starting point for reading the CSV
csv_start_row = 7  # Row index in CSV to start reading from (0-based index - Does not include headers)
csv_start_col = 0  # Column index in CSV to start reading from (0-based index)
excel_start_row = 2  # Row in Excel to start writing (1-based index)
excel_start_col = 2  # Column in Excel to start writing (1-based index)


# Filter the CSV data based on the starting row and column
data_to_write = data.iloc[csv_start_row:, csv_start_col:]  # Slicing CSV data
# Open the Excel workbook using xlwings
ws = wb.sheets["dk_list"]
# Define the starting point for writing to Excel
ws.range((excel_start_row, excel_start_col)).value = data_to_write.values

#---------------------------------------------------------------------------------------#

# **2. Write `DKSalaries.csv` to "sog_projections" sheet with different settings**
# Modify settings for the new sheet
csv_start_row = 7  # New row index for "sog_projections"
csv_start_col = 0  # New column index for "sog_projections"
excel_start_row = 2
excel_start_col = 2

# Filter the CSV data again with new settings
data_to_write = data.iloc[csv_start_row:, csv_start_col:]

# Write to the "sog_projections" sheet
ws = wb.sheets["sog_projections"]
ws.range((excel_start_row, excel_start_col)).value = data_to_write.values

# **3. Write `DARKO_player_talent_2024-12-02.csv` to "darko" sheet**
csv_file_darko = "DARKO_player_talent_2024-12-02.csv"
data_darko = pd.read_csv(csv_file_darko)

# Define settings for "darko" sheet
csv_start_row = 0  # Starting row in the CSV
csv_start_col = 0  # Starting column in the CSV
excel_start_row = 2  # Starting row in Excel
excel_start_col = 1  # Starting column in Excel

# Filter the CSV data
data_to_write_darko = data_darko.iloc[csv_start_row:, csv_start_col:]

# Write to the "darko" sheet
ws = wb.sheets["darko"]
ws.range((excel_start_row, excel_start_col)).value = data_to_write_darko.values



# **3. Write `advanced.csv` to "advanced" sheet**
csv_file_advanced = "advanced.csv"
data_advanced = pd.read_csv(csv_file_advanced)

# Define settings for "advanced" sheet
csv_start_row = 0  # Starting row in the CSV
csv_start_col = 1  # Starting column in the CSV
excel_start_row = 2  # Starting row in Excel
excel_start_col = 2  # Starting column in Excel

# Filter the CSV data
data_to_write_advanced = data_advanced.iloc[csv_start_row:, csv_start_col:]

# Write to the "advanced" sheet
ws = wb.sheets["team_ratings"]
ws.range((excel_start_row, excel_start_col)).value = data_to_write_advanced.values

# Save and close the workbook
wb.save()
wb.close()
app.quit()












'''
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


# Define the current script folder and sibling folder
#current_folder = Path(__file__).parent  # Current script directory (src)
#target_folder = current_folder.parent / "dk-data"  # Sibling folder (dk-data)

# Define file paths
#excel_file = target_folder / "your_excel_file.xlsx"
#csv_file = target_folder / "your_csv_file.csv"

# Load the workbook and worksheet
wb = load_workbook("nba_dfs_optimized.xlsm", keep_vba=True)
ws = wb["dk_list"]

# Load the CSV data
data = pd.read_csv("DKSalaries.csv")

# Define the starting point for reading the CSV
csv_start_row = 0  # Row index in CSV to start reading from (0-based index - Does not include headers)
csv_start_col = 1  # Column index in CSV to start reading from (0-based index)

# Define the starting point for writing to Excel
excel_start_row = 2  # Row in Excel to start writing (1-based index)
excel_start_col = 2  # Column in Excel to start writing (1-based index)

# Filter the CSV data based on the starting row and column
data_to_write = data.iloc[csv_start_row:, csv_start_col:]  # Slicing CSV data

# Write the filtered data to the Excel sheet
for i, row in enumerate(data_to_write.itertuples(index=False), start=excel_start_row):
    for j, value in enumerate(row, start=excel_start_col):
        print(value)
        ws.cell(row=i, column=j, value=value)

# Save the workbook
wb.save("nba_dfs_optimized.xlsm")
'''
